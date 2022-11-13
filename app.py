import tkinter as tk
from tkinter import ttk
import openpyxl as op

from matplotlib.transforms import composite_transform_factory

window = tk.Tk()
LabelFont  = ("Helevetice", 18)
titleLabelFont  = ("Helevetice", 18)
wb=op.load_workbook('Excel_Sheet.xlsx')


def row_acquisition():
    sheet = wb[date]
    maxRow = sheet.max_row + 1
    maxClm = sheet.max_column + 1
    for j in range(1,maxClm):
        if j == 2:
            for i in reversed(range(1,maxRow)):
                if sheet.cell(row=i, column=j).value != None:
                    return i

def changePage():
    page1()

def cp_return(event):
    page1()

def cp_date():
    global date
    global fieldnames
    global w
    d=str(date_input.get())
    date = d
    print(type(date))
    #print(date)
    sheet_copy = wb.copy_worksheet(wb['Template'])
    ws = wb.worksheets[-1]
    ws.title=date
    sheet = wb[date]
    sheet['E3']= d[:4] + '年' + d[4:6] + '月' + d[6:8] + '日'
    wb.save('Excel_Sheet.xlsx')
    page2()

def cp_date_return(event):
    cp_date()
    

def cp_name():
    global name
    name=str(name_input.get())
    #print(name)
    page3()

def cp_name_return(event):
    cp_name()
    
def cp_codename():
    global code
    global number
    c=str(code_input.get())
    code=c.upper()
    number=int(number_input.get())
    # print(code)
    # print(number)
    page4()

def cp_codename_return(event): 
    cp_codename()
    
def cp_capacity():
    global capacity
    capacity=float(capacity_input.get())
    # print(capacity)
    page5()

def cp_capacity_return(event):
    cp_capacity()
    
def cp_temp():
    global temp
    temp=int(temp_input.get())
    # print(temp)
    page6()

def cp_temp_return(event):
    cp_temp()

def cp_amount():
    global amount
    amount=float(amount_input.get())
    # print(type(amount))
    # print(amount)
    page7()

def cp_amount_return(event):
    cp_amount()
    
def cp_increment():
    global increment
    increment=float(increment_input.get())
    # print(type(increment))
    # print(increment)
    calculation()
    page8()

def cp_increment_return(event):
    cp_increment()
    
def fin():
    window.destroy()
    sheet = wb[date]
    r = str(row_acquisition() + 1)
    sheet['B' + r] = name
    sheet['C' + r] = code
    sheet['D' + r] = number
    sheet['E' + r] = total_increase
    sheet['F' + r] = increment
    sheet['G' + r] = pri
    wb.save('Excel_Sheet.xlsx')

def fin_return(event):
    fin()
        
def add():
    sheet = wb[date]
    r = str(row_acquisition() + 1)
    sheet['B' + r] = name
    sheet['C' + r] = code
    sheet['D' + r] = number
    sheet['E' + r] = total_increase
    sheet['F' + r] = increment
    sheet['G' + r] = pri
    wb.save('Excel_Sheet.xlsx')
    page2()
    
    
def calculation():
    global result
    global total_increase
    global permanent_rate_increase
    global pri
    f_list=[0.000505,0.000502,0.000498,0.000495,0.000492,0.000489,0.000486,0.000484,0.000482,0.000479,0.000477,0.000475,0.000474,0.000472,0.000470,0.000469,0.000467,0.000466,0.000465,0.000464,0.000462,0.000461,0.000460,0.000459,0.000458,0.000457,0.000456,0.000455,0.000455,0.000454,0.000453,0.000452,0.000451,0.000449,0.000448,0.000447,0.000446,0.000445,0.000443,0.000442]
    b=6
    p=3
    c=capacity
    a=amount
    i=increment
    t=temp
    v=c*1000
    num=t-1
    beta=f_list[num]
    total_increase=(a-b)-((a-b)+v)*p*beta
    permanent_rate_increase=(i/total_increase)*100
    pri=round(permanent_rate_increase,2)
    if pri < 0:
        pri = 0
    # print('リスト番号：%d'%num)
    # print('全入水量：%d'%a)
    # print('内容積：%d'%v)
    # print('耐圧試験圧力：%d'%p)
    # print('水温：%d'%t)
    # print(beta)
    # print('全増加量：%f'%total_increase)
    # print('恒久増加率：%f'%permanent_rate_increase)
    # print('率：%f'%pri)
    if permanent_rate_increase < 10:
        result='合格'
    else:
        result='不合格'
    
#input date    
def page1():
    global date
    global date_input
    Page1 = tk.Frame(window)
    spaceLabel3 = [tk.Label(Page1, text="") for column in range(5)]
    titleLabel=ttk.Label(Page1, text="日付を入力してください。(例:2020年1月1日→20200101)", font=titleLabelFont)

    for index in range(5):
        spaceLabel3[index].pack()
   
    titleLabel.pack()

    frame = ttk.Frame(Page1)

    frame.pack()

    spaceLabel3 = [tk.Label(frame, text="") for column in range(3)]

    dateLabelFont  = ("Helevetice", 18)
    dateLabel      = ttk.Label(frame, text="日付：", font=dateLabelFont)

    for index in range(3):
        spaceLabel3[index].grid(row=index, column=0)
    dateLabel.grid(row=4, column=0)
    date = tk.Entry()
    dateEntry = ttk.Entry(frame, textvariable=date, width=30,font= LabelFont)
    date_input=dateEntry
    dateEntry.grid(row=4, column=1) 
    okButton = ttk.Button(frame, text="  次へ  ", command=lambda : cp_date())
    okButton.grid(row=5, column=3)
    Page1.grid(row=0, column=0, sticky="nsew")
    window.bind('<Return>',cp_date_return)

#inpupt customer name
def page2():
    global name
    global name_input
    Page2 = tk.Frame(window)
    spaceLabel4 = [tk.Label(Page2, text="") for column in range(5)]
    titleLabel      =\
       ttk.Label(Page2, text="客先名を入力してください。", font=titleLabelFont)
  
    for index in range(5):
        spaceLabel4[index].pack()

    titleLabel.pack()

    frame = ttk.Frame(Page2)

    frame.pack()

    spaceLabel4 = [tk.Label(frame, text="") for column in range(3)]

    nameLabelFont  = ("Helevetice", 18)
    nameLabel      = ttk.Label(frame, text="客先名：", font=nameLabelFont)

    for index in range(3):
        spaceLabel4[index].grid(row=index, column=0) 
    nameLabel.grid(row=4, column=0)
    name = tk.Entry()
    nameEntry = ttk.Entry(frame, textvariable=name, width=30,font=LabelFont)
    name_input=nameEntry
    nameEntry.grid(row=4, column=1)

    okButton = ttk.Button(frame, text="  次へ  ", command=lambda : cp_name())

    okButton.grid(row=5, column=3)

    Page2.grid(row=0, column=0, sticky="nsew")

    window.bind('<Return>',cp_name_return)
    
#input container code and number
def page3():
    global code
    global code_input
    global number
    global number_input
    Page3 = tk.Frame(window)
    spaceLabel5 = [tk.Label(Page3, text="") for column in range(5)]
    titleLabel      =\
       ttk.Label(Page3, text="容器記号番号を入力してください。", font=titleLabelFont)
  
    for index in range(5):
        spaceLabel5[index].pack()

    titleLabel.pack()

    frame = ttk.Frame(Page3)

    frame.pack()

    spaceLabel6 = [tk.Label(frame, text="") for column in range(3)]
    spaceLabel7 = [tk.Label(frame, text="") for column in range(3)]

  
    codeLabelFont  = ("Helevetice", 18)
    codeLabel      = ttk.Label(frame, text="記号：", font=codeLabelFont)
    for index in range(3):
        spaceLabel6[index].grid(row=index, column=0)   
    codeLabel.grid(row=2, column=0)
    code = tk.Entry()
    codeEntry = ttk.Entry(frame, textvariable=code, width=20,font=LabelFont)
    codeEntry.grid(row=2, column=1)
    code_input=codeEntry
    
    numberLabelFont  = ("Helevetice", 18)
    numberLabel      = ttk.Label(frame, text="番号：", font=numberLabelFont)
    for index in range(3):
        spaceLabel7[index].grid(row=index, column=0)  
    numberLabel.grid(row=4, column=0)
    number = tk.Entry()
    numberEntry = ttk.Entry(frame, textvariable=number, width=20,font=LabelFont)
    numberEntry.grid(row=4, column=1)
    number_input=numberEntry

    okButton = ttk.Button(frame, text="  次へ  ", command=lambda : cp_codename())
    okButton.grid(row=5, column=3)
    Page3.grid(row=0, column=0, sticky="nsew")

    window.bind('<Return>',cp_codename_return)

#input capacity
def page4():
    global capacity
    global capacity_input
    Page4 = tk.Frame(window)
    spaceLabel8 = [tk.Label(Page4, text="") for column in range(5)]
    titleLabel      =\
       ttk.Label(Page4, text="内容積を入力してください。(L)", font=titleLabelFont)
  
    for index in range(5):
        spaceLabel8[index].pack()

    titleLabel.pack()
    frame = ttk.Frame(Page4)
    frame.pack()
    spaceLabel9 = [tk.Label(frame, text="") for column in range(3)]
    capacityLabelFont  = ("Helevetice", 18)
    capacityLabel      = ttk.Label(frame, text="内容積：", font=capacityLabelFont)

    for index in range(3):
        spaceLabel9[index].grid(row=index, column=0)
    
    capacityLabel.grid(row=4, column=0)
    capacity = tk.Entry()
    capacityEntry = ttk.Entry(frame, textvariable=capacity, width=30,font=LabelFont)
    capacity_input=capacityEntry
    capacityEntry.grid(row=4, column=1)

    okButton = ttk.Button(frame, text="  次へ  ", command=lambda : cp_capacity())
    okButton.grid(row=5, column=3)

    Page4.grid(row=0, column=0, sticky="nsew")

    window.bind('<Return>',cp_capacity_return)

#input water temperature
def page5():
    global temp
    global temp_input
    Page5 = tk.Frame(window)
    spaceLabel10 = [tk.Label(Page5, text="") for column in range(5)]
    titleLabel      =\
       ttk.Label(Page5, text="水温を入力してください。(℃)", font=titleLabelFont)
  
    for index in range(5):
        spaceLabel10[index].pack()

    titleLabel.pack()

    frame = ttk.Frame(Page5)

    frame.pack()

    spaceLabel11 = [tk.Label(frame, text="") for column in range(3)]
    tempLabelFont  = ("Helevetice", 18)
    tempLabel      = ttk.Label(frame, text="水温：", font=tempLabelFont)

    for index in range(3):
        spaceLabel11[index].grid(row=index, column=0)
    
    tempLabel.grid(row=4, column=0)
    temp = tk.Entry()
    tempEntry = ttk.Entry(frame, textvariable=temp, width=30,font=LabelFont)
    temp_input=tempEntry
    tempEntry.grid(row=4, column=1)

    okButton = ttk.Button(frame, text="  次へ  ", command=lambda : cp_temp())
    okButton.grid(row=5, column=3)

    Page5.grid(row=0, column=0, sticky="nsew")

    window.bind('<Return>',cp_temp_return)
    
#input total water inflow
def page6():
    global amount
    global amount_input
    Page6 = tk.Frame(window)
    spaceLabel1 = [tk.Label(Page6, text="") for column in range(5)]
    titleLabel      =\
       ttk.Label(Page6, text="全圧入水量を入力してください。(㎤)", font=titleLabelFont)
  
    for index in range(5):
        spaceLabel1[index].pack()

    titleLabel.pack()

    frame = ttk.Frame(Page6)

    frame.pack()

    spaceLabel2 = [tk.Label(frame, text="") for column in range(3)]

  
    amountLabelFont  = ("Helevetice", 18)
    amountLabel      = ttk.Label(frame, text="全圧入水量：", font=amountLabelFont)

    for index in range(3):
        spaceLabel2[index].grid(row=index, column=0)
    
    amountLabel.grid(row=4, column=0)

    
    amount = tk.Entry()
    amountEntry = ttk.Entry(frame, textvariable=amount, width=30,font=LabelFont)
    amount_input=amountEntry
 
    amountEntry.grid(row=4, column=1)


    okButton = ttk.Button(frame, text="  次へ  ", command=lambda : cp_amount())
    okButton.grid(row=5, column=3)

    Page6.grid(row=0, column=0, sticky="nsew")

    window.bind('<Return>',cp_amount_return)

#input permanent increase
def page7():
    global increment
    global increment_input
    Page7 = tk.Frame(window)
    spaceLabel1 = [tk.Label(Page7, text="") for column in range(5)]
    titleLabel      =ttk.Label(Page7, text="恒久増加量を入力してください。(㎤)", font=titleLabelFont)
  
    for index in range(5):
        spaceLabel1[index].pack()

    titleLabel.pack()

    frame = ttk.Frame(Page7)

    frame.pack()

    spaceLabel2 = [tk.Label(frame, text="") for column in range(3)]
    incrementLabelFont  = ("Helevetice", 18)
    incrementLabel      = ttk.Label(frame, text="恒久増加量：", font=incrementLabelFont)

    for index in range(3):
        spaceLabel2[index].grid(row=index, column=0)
    
    incrementLabel.grid(row=4, column=0)
    increment = tk.Entry()
    incrementEntry = ttk.Entry(frame, textvariable=increment, width=30,font=LabelFont)
    increment_input=incrementEntry
 
    incrementEntry.grid(row=4, column=1)

    # return_Button = ttk.Button(frame, text="  戻る  ", command=lambda : cp_temp())
    # return_Button.grid(row=5, column=2)

    okButton = ttk.Button(frame, text="  次へ  ", command=lambda : cp_increment())
    okButton.grid(row=5, column=3)

    Page7.grid(row=0, column=0, sticky="nsew")

    window.bind('<Return>',cp_increment_return)

#show result
def page8():
    Page8 = tk.Frame(window)
    spaceLabel16 = [tk.Label(Page8, text="") for column in range(5)]
    titleLabel      =\
       ttk.Label(Page8, text=" ", font=titleLabelFont)
  
    for index in range(5):
        spaceLabel16[index].pack()

    titleLabel.pack()
    frame = ttk.Frame(Page8)

    frame.pack()

    spaceLabel17 = [tk.Label(frame, text="") for column in range(3)]
    spaceLabel18 = [tk.Label(frame, text="") for column in range(3)]

  
    rateLabelFont  = ("Helevetice", 18)
    rateLabel      = ttk.Label(frame, text="恒久増加率：", font=rateLabelFont)
    for index in range(3):
        spaceLabel17[index].grid(row=index, column=0)
    rateLabel.grid(row=2, column=0)
    if pri > 0:
        output1=ttk.Label(frame, text=pri,font=LabelFont)
    else:
        output1=ttk.Label(frame, text=0,font=LabelFont)

    output1.grid(row=2,column=2)

    per=ttk.Label(frame, text='%',font=LabelFont)
    per.grid(row=2,column=3)


    resultLabelFont  = ("Helevetice", 18)
    resultLabel      = ttk.Label(frame, text="合否：", font=resultLabelFont)
    for index in range(3):
        spaceLabel18[index].grid(row=index, column=0)   
    resultLabel.grid(row=4, column=0)
    output2=ttk.Label(frame, text=result,font=LabelFont)
    output2.grid(row=4,column=1)
    

    return_Button = ttk.Button(frame, text="  戻る  ", command=lambda : cp_amount())
    return_Button.grid(row=5, column=2)
    
    next_Button = ttk.Button(frame, text="  次へ  ", command=lambda : add())
    next_Button.grid(row=5, column=3)
    
    finish_Button = ttk.Button(frame, text="  完了  ", command=fin)
    finish_Button.grid(row=7, column=3)

    Page8.grid(row=0, column=0, sticky="nsew")

    window.bind('<Return>',fin_return)

def main() -> None:
    window.title("容器再検査")

    window.geometry("800x600")

    window.grid_rowconfigure(0, weight=1)
    window.grid_columnconfigure(0, weight=1)

    startPage = tk.Frame(window)

    spaceLabel1 = [tk.Label(startPage, text="") for column in range(10)]
    spaceLabel2 = [tk.Label(startPage, text="") for column in range(3)]

    titleLabelFont  = ("Helevetice", 32, "bold")
    titleLabel      = ttk.Label(startPage, text="容器再検査", font=titleLabelFont)

    for index in range(10):
        spaceLabel1[index].pack()
        
    titleLabel.pack()
    startButton =\
    ttk.Button(startPage, text="           開始           ", command=lambda : changePage())
    window.bind('<Return>',cp_return)
    for index in range(3):
        spaceLabel2[index].pack()
        
    startButton.pack()
    startPage.grid(row=0, column=0, sticky="nsew")
    startPage.tkraise()
    window.mainloop()

if __name__ == "__main__":
    main()